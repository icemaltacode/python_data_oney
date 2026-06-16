"""
A Summary Row
=============

Reports usually end with a summary. You'll add the average height two different
ways so you can see the difference:

  1. As an Excel FORMULA (a string starting with "="). Excel works it out when
     the file is opened, and it recalculates if the data changes.
  2. As a VALUE computed in Python with NumPy. It's frozen the moment you write
     it.

The dataset has 1015 players, so the data sits in rows 2 to 1016 and your
summary goes in row 1017.

The workbook is filled with data for you. Add a label, the formula, and the
NumPy value, then save.
"""

# region setup
import pandas as pd
import numpy as np
from openpyxl import Workbook

data = pd.read_csv("Module 1/data/baseball.csv")[["Name", "Height", "Weight", "Age"]]

wb = Workbook()
ws = wb.active
ws.title = "Players"
ws.append(["Name", "Height", "Weight", "Age"])
for row in data.itertuples(index=False):
    ws.append(row)
# endregion

## ---- START HERE ----

# Write the label "Average" into cell A1017


# Write an Excel formula into B1017 that averages the Height column (B2:B1016)
ws["B1017"] =

# Compute the average height in Python with NumPy, then write it to C1017
avg_height =
ws["C1017"] = avg_height

# Save the workbook as "report.xlsx"

