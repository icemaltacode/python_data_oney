"""
Make it Presentable
====================

A report nobody can read is a report nobody uses. In this final exercise you'll
style the header row and add a chart.

The workbook is filled with data for you. Your job:

  1. Make the header row (row 1) bold with white text on a dark blue fill.
  2. Widen column A so the names fit.
  3. Add a scatter chart of Height vs. Weight - the same relationship you found
     correlated at 0.58 back in Project 3.

Then save. Open report.xlsx and admire your work.
"""

# region setup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series

data = pd.read_csv("Module 4/data/baseball.csv")[["Name", "Height", "Weight", "Age"]]

wb = Workbook()
ws = wb.active
ws.title = "Players"
ws.append(["Name", "Height", "Weight", "Age"])
for row in data.itertuples(index=False):
    ws.append(row)
# endregion

## ---- START HERE ----

# Style every cell in the header row (ws[1]): bold + white font, dark blue fill.
# Hint: Font(bold=True, color="FFFFFF") and PatternFill("solid", fgColor="1F3864")
for cell in ws[1]:
    cell.font = ____
    cell.fill = ____

# Widen column "A" to 22
ws.column_dimensions["A"].width = ____

# Build a scatter chart of Height (column 2) vs Weight (column 3).
# Data is in rows 2 to 1016.
chart = ScatterChart()
chart.title = "Height vs. Weight"

x_values = Reference(ws, min_col=2, min_row=2, max_row=1016)
y_values = Reference(ws, min_col=3, min_row=2, max_row=1016)

# Create a Series from (y_values, x_values) and append it to chart.series


# Add the chart to the worksheet anchored at cell "F2"


# Save the workbook as "report.xlsx"

